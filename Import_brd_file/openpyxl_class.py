import openpyxl
import sys
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
import numpy as np

class PyEasyExcel():    
    """A utility to make it easier to get at Excel.    Remembering  
    to save the data is your problem, as is    error handling.  
    Operates on one workbook at a time."""    
    def __init__(self, filename=None):  #打开文件或者新建文件（如果不存在的话）    
        if filename:    
            self.filename = filename    
            self.wb = openpyxl.load_workbook(filename)  
        else:    
            self.wb = openpyxl.Workbook()    
            self.filename = 'Summary.xlsx'  

    def save(self, newfilename=None):  #保存文件  
        if newfilename:    
            self.filename = newfilename    
            self.wb.save(self.filename)    
        else:    
            self.wb.save(self.filename)        
    def close(self):  #关闭文件  
        self.wb.close()    
        del self 
    def getsheetname(self):
        sheet = self.wb.get_sheet_names()
        return sheet     
    def sheetAdd(self,sheetname=None):
        if sheetname:
            self.wb.create_sheet(sheetname)
        else:
            self.wb.create_sheet()
    def getCell(self, sheetname, row, col):  #获取单元格的数据  
        "Get value of one cell"    
        sht = self.wb.get_sheet_by_name(sheetname)    
        return sht.cell(row=row,column=col).value   
    def getMaxRow(self, sheetname):  #获取sheet使用的最大行数  
        "Get value of max row"    
        sht = self.wb.get_sheet_by_name(sheetname)    
        return sht.max_row     
    def getMaxCol(self, sheetname):  #获取sheet使用的最大列数  
        "Get value of max columns"    
        sht = self.wb.get_sheet_by_name(sheetname)   
        return sht.max_column                    
    def setCell(self, sheetname, r, c, value):  #设置单元格的数据  
        "set value of one cell"    
        sht = self.wb.get_sheet_by_name(sheetname)    
        sht.cell(row=r, column=c).value = value  
    def setCellformat(self, sheetname, row, col):  #设置单元格的格式  
        "set value of one cell"    
        sht = self.wb.get_sheet_by_name(sheetname)    
        # sht.cell(row=row, col=col).Font.Size = 15#字体大小  
        # sht.Cells(row, col).Font.Bold = True#是否黑体  
        # sht.Cells(row, col).Name = "Arial"#字体类型  
        # sht.Cells(row, col).Interior.ColorIndex = 3#表格背景  
        # #sht.Range("A1").Borders.LineStyle = xlDouble  
        # sht.Cells(row, col).BorderAround(1,4)#表格边框  
        # sht.Rows(3).RowHeight = 30#行高  
        # sht.Cells(row, col).HorizontalAlignment = -4131 #水平居中xlCenter  
        # sht.Cells(row, col).VerticalAlignment = -4160 #  
    def deleteRow(self, sheetname, row):  
        sht = self.wb.get_sheet_by_name(sheetname)  
        sht.delete_rows(row)#删除行  
    def deleteCol(self, sheetname, col):  
        sht = self.wb.get_sheet_by_name(sheetname)  
        sht.delete_cols(col)#删除列
    def getRange(self, sheetname, row1, col1, row2, col2):  #获得一块区域的数据，返回为一个二维元组  
        "return a 2d array (i.e. tuple of tuples)"    
        sheet = self.wb.get_sheet_by_name(sheetname)
        return (tuple(tuple(sheet.cell(row=row1+i, column=col1+j).value for j in range(col2-col1+1)) for i in range(row2-row1+1)))   
    def setRange(self, data_array, sheetname, row1, col1, row2, col2):  #根据一个二维元组，写入一块区域数据  
        "write a 2d array (i.e. tuple of tuples)"    
        sheet = self.wb.get_sheet_by_name(sheetname) 
        if row2 == row1:
            for j in range(col2-col1+1):
                sheet.cell(row=row1, column=col1+j).value = data_array[j]
        else:
            for i in range(row2-row1+1):
                for j in range(col2-col1+1):
                    sheet.cell(row=row1+i, column=col1+j).value = data_array[i][j]  
                   
    def addPicture(self, sheetname, pictureName, Left, Top, Width, Height):  #插入图片  
        "Insert a picture in sheet"    
        sheet = self.wb.get_sheet_by_name(sheetname)   
        img = Image(pictureName) 
        img.width, img.height=Width, Height
        sheet.add_image(img,'D1')    

    def cpSheet(self, source_sheet, new_sheet):  #复制工作表  
        "copy sheet"    
        sht = self.copy_worksheet(wb[source_sheet])   
        sht.title = new_sheet
        

    def insertRow(self,sheetname,row):
        sheet = self.wb.get_sheet_by_name(sheetname)
        sheet.insert_rows(row)
    def insertCol(self,sheetname,col):
        sheet = self.wb.get_sheet_by_name(sheetname)
        sheet.insert_cols(col)       
